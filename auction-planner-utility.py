# !/usr/bin/env python
"""
Jeff's Lien Sale Event Tracker
This helps with self storage planning for unit auctions. It requires a specific Excel export from Sitelink operational
self storage software. The code is useful to me for reference because it contains methods for reading and handling
files, date math, applications of the PySimpleGui framework, and it contains simple methods for accessing values of
nested dictionaries.  - Jeffrey Neil Willits
"""

import PySimpleGUI as sg
import json
import os
import sys
import openpyxl
from datetime import datetime
from dateutil.parser import *

sg.ChangeLookAndFeel('Dark')
sg.SetOptions(icon='aplanner_icon.ico', element_padding=(6, 6), font=('verdana', 9), text_color='#32CD32',
              background_color='#1E1E1E', text_element_background_color='#1E1E1E', button_color=('#FFFFFF', '#2F2F2F'))
menu_def = [['Setup', ['Input File', 'Output File', 'Backup Folder']],
            ['Help', 'About...']]

layout = [
    [sg.Menu(menu_def, tearoff=False, pad=(20, 1))],
    [sg.T('')],
    [sg.T('')],
    [sg.T('', size=(200, 1), key='_INPUT_FILE_PATH_'), ],
    [sg.T('')],
    [sg.T('', size=(200, 1), key='_OUTPUT_FILE_PATH_'), ],
    [sg.T('')],
    [sg.T('', size=(200, 1), key='_BACKUP_FOLDER_PATH_'), ],
    [sg.T('')],
    [sg.T('')],
    [sg.Button('', visible=False, size=(12, 1), ), ],
    [sg.Button('Run Update', visible=True, size=(12, 1), )],
    [sg.Button('Cancel', visible=True, size=(12, 1), ), ]]


def define_file(f_use_str):
    prompt_str = f'\nIdentify the {f_use_str} file...\n\n'
    sg.Popup('', prompt_str, background_color='#183a3e', text_color='#ffffff', keep_on_top=True)
    if len(sys.argv) == 1:
        event, (file_path,) = sg.Window('My Script').Layout([[sg.Text('Document to open')],
                                                             [sg.In(size=(100, 10)), sg.FileBrowse()],
                                                             [sg.CloseButton('Open'), sg.CloseButton('Cancel')]]).Read()
    else:
        file_path = sys.argv[1]
    if not file_path:
        sg.Popup("Cancel", "No file path was supplied.")
        raise SystemExit("Cancelling - no file path was supplied.")
    return file_path


def backup_folder(f_use_str):
    prompt_str = f'\nIdentify the {f_use_str} file backup folder location...\n\n'
    sg.Popup('', prompt_str, background_color='#183a3e', text_color='#ffffff', keep_on_top=True)
    if len(sys.argv) == 1:
        event, (file_path,) = sg.Window('My Script').Layout([[sg.Text('Backup files destination')],
                                                             [sg.In(size=(100, 10)), sg.FileBrowse()],
                                                             [sg.CloseButton('Open'), sg.CloseButton('Cancel')]]).Read()
    else:
        file_path = sys.argv[1]
    if not file_path:
        sg.Popup("Cancel", "No folder was supplied.")
        raise SystemExit("Cancelling - no file folder path was supplied.")
    return file_path


def read_files():
    if os.path.isfile('filepath.json'):
        with open('filepath.json') as f_obj:
            return json.load(f_obj)
    else:
        input_file_path = define_file('input')
        output_file_path = define_file('output')
        backup_folder_path = backup_folder('backup')
        return {'input': input_file_path, 'output': output_file_path, 'backup': backup_folder_path}


def write_data(file_path_dict_pass):
    with open('filepath.json', 'w') as f_obj:
        json.dump(file_path_dict_pass, f_obj)


def load_data(ws_pass, unit_col_pass, head_row_num_pass):
    """ Iterate through the columns and return a list. """
    ws_list = []
    # ToDo: Does this know to stop with empty rows?
    for row in ws_pass[unit_col_pass]:
        if row.value != 'x':
            ws_list.append(row.value)
    for i in range(0, head_row_num_pass):
        ws_list.pop(0)
    return ws_list


def populate_data_objects(f_str, unit_col, date_col, head_row_num, sheet_name):
    """ Extract data from the unit and paid-thru spreadsheet columns and insert it in lists. Then populate a dictionary
        with the data. """
    wb = openpyxl.load_workbook(f_str)
    ws = wb[sheet_name]
    ws_dict = {}
    ws_unit_list = load_data(ws, unit_col, head_row_num)
    ws_date_list = load_data(ws, date_col, head_row_num)
    for i in range(0, len(ws_unit_list)):
        ws_dict[ws_unit_list[i]] = ws_date_list[i]
    return wb, ws, ws_unit_list, ws_dict


def core_tasks(ws1_file, ws2_file, backup_folder_str):
    """ This is the main operational function. It starts by getting data from both spreadsheets to lists
        and dictionaries. """

    wb1, ws1, ws1_unit_list, ws1_dict = populate_data_objects(ws1_file, 'C', 'O', 1, 'Sheet1')
    wb2, ws2, ws2_unit_list, ws2_dict = populate_data_objects(ws2_file, 'B', 'D', 6, 'Lien Sale Tracker')
    wb2.save(backup_folder_str[:-5] + datetime.now().strftime('%Y-%m-%d %H%M%S') + '.xlsx')

    # Determine units to delete from the output spreadsheet (paid accounts). The paid_unit_list will contain units
    # listed on the output that are not in the input. The paid_unit_list will also contain units that have paid-thru
    # dates which are different than dates on the input spreadsheet.

    sheet_row = 7
    i = 0
    while True:
        if i == len(ws2_unit_list):
            break
        if ws2_unit_list[i] not in ws1_unit_list:
            del ws2_unit_list[i]
            ws2.delete_rows(sheet_row)
        sheet_row += 1
        i += 1

    wb2.save(ws2_file)
    wb2.close()
    wb2, ws2, ws2_unit_list, ws2_dict = populate_data_objects(ws2_file, 'B', 'D', 6, 'Lien Sale Tracker')

    sheet_row = 7
    i = 0
    while True:
        if i == len(ws2_unit_list):
            break
        if ws2_unit_list[i] in ws2_dict and ws2_unit_list[i] in ws1_dict:
            if str(parse(str(ws2_dict[ws2_unit_list[i]])).date()) != str(parse(str(ws1_dict[ws2_unit_list[i]])).date()):
                del ws2_unit_list[i]
                ws2.delete_rows(sheet_row)
        sheet_row += 1
        i += 1

    # Add new past-due accounts.
    wb2.save(ws2_file)
    wb2.close()
    wb2, ws2, ws2_unit_list, ws2_dict = populate_data_objects(ws2_file, 'B', 'D', 6, 'Lien Sale Tracker')
    i = 0
    sheet_row = 7 + len(ws2_unit_list)
    while True:
        if i == len(ws1_unit_list):
            break
        if ws1_unit_list[i] not in ws2_unit_list:
            ws2.cell(row=sheet_row, column=2).value = ws1_unit_list[i]
            ws2.cell(row=sheet_row, column=4).value = ws1_dict[ws1_unit_list[i]]
            sheet_row += 1
        i += 1

    # Make dates match input.
    wb2.save(ws2_file)
    wb2.close()
    wb2, ws2, ws2_unit_list, ws2_dict = populate_data_objects(ws2_file, 'B', 'D', 6, 'Lien Sale Tracker')
    i = 0
    sheet_row = 7
    while True:
        if i == len(ws2_unit_list):
            break

        if ws2_unit_list[i] in ws1_dict:
            ws2.cell(row=sheet_row, column=4).value = ws1_dict[ws2_unit_list[i]]
        sheet_row += 1
        i += 1

    wb2.save(ws2_file)
    sg.Popup('', 'Processing is complete and the spreadsheet is updated.', no_titlebar=True, keep_on_top=True,
             grab_anywhere=True, background_color='#183a3e', text_color='#ffffff')
    wb2.close()
    file_path_dict['input'] = ws1_file_path
    file_path_dict['output'] = ws2_file_path
    write_data(file_path_dict)
    quit()


if __name__ == '__main__':
    file_path_dict = read_files()
    ws1_file_path = file_path_dict['input']
    ws2_file_path = file_path_dict['output']
    backup_folder_str = file_path_dict['backup']
    path_length = max(len(ws1_file_path), len(ws2_file_path), len(backup_folder_str))
    window = sg.Window(" Jeff's Auction Planner Utility", size=(path_length + 700, 400), default_element_size=(30, 1),
                       grab_anywhere=False, background_color='#1E1E1E', auto_size_text=False,
                       auto_size_buttons=False).Layout(layout).Finalize()
    window.Element('_INPUT_FILE_PATH_').Update(f'Input file path: {ws1_file_path}')
    window.Element('_OUTPUT_FILE_PATH_').Update(f'Output file path: {ws2_file_path}')
    window.Element('_BACKUP_FOLDER_PATH_').Update(f'Backup file folder path: {backup_folder_str}')
    while True:
        event, values = window.Read(timeout=10)
        if event is None or event == 'Exit':
            break
        else:
            if event == 'About...':
                sg.Popup("Jeff's Auction Tracker Utility", 'This is a utility to help track legally required events',
                         'that lead to a self storage unit auction. It assures',
                         'the time interval between events adheres to requirements',
                         'of the Hawaii lien laws. When the program starts, it',
                         'searches for a specific Excel file that can be exported',
                         'from Sitelink. Sitelink is a popular operating software',
                         'for self storage. When event dates are marked, this',
                         'information is stored in a separate file. Obsolete records',
                         'are deleted automatically when the Excel file is read.\n',
                         'The program is not copyrighted and it is free for use.',
                         'Python source code for this is available in my public',
                         'GitHub repository.\n',
                         'Version 1.1 finished May 12, 2019.\n\n'
                         'Jeffrey Neil Willits', 'W: jnwillits.com\n', no_titlebar=True, keep_on_top=True,
                         grab_anywhere=True, background_color='#183a3e', text_color='#ffffff')
            elif event == 'Input File':
                ws1_file_path = define_file('input')
                window.Element('_INPUT_FILE_PATH_').Update(f'Input file path: {ws1_file_path}')
            elif event == 'Output File':
                ws2_file_path = define_file('output')
                window.Element('_OUTPUT_FILE_PATH_').Update(f'Output file path: {ws2_file_path}')
            elif event == 'Backup Folder':
                backup_folder_str = backup_folder('backup')
                window.Element('_BACKUP_FOLDER_PATH_').Update(f'Backup file destination: {ws2_file_path}')
            elif event == 'Run Update':
                core_tasks(ws1_file_path, ws2_file_path, backup_folder_str)
            elif event == 'Cancel':
                exit()
                window.Close()

    file_path_dict['input'] = ws1_file_path
    file_path_dict['output'] = ws2_file_path
    write_data(file_path_dict)
    window.Close()
